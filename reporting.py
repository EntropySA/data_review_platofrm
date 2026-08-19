"""Excel reporting adapter for completed reviews."""

from io import BytesIO
from typing import Mapping, Sequence


EXPORT_COLUMNS = ("instruction", "question", "output", "pass/fail", "notes")
FAILURE_COLUMNS = ("source_id", "instruction", "input", "output", "notes", "check", "row_hash")
UNCHECKED_COLUMNS = ("source_id", "reason")

FAILURE_SHEET = "الأخطاء"
UNCHECKED_SHEET = "غير مفحوص"


def _style_sheet(worksheet, widths) -> None:
    """Header band, column widths, wrapped body, frozen header and a filter."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="164E63")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")

    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 24


def create_excel_export(rows: Sequence[Mapping[str, str]]) -> bytes:
    # Imported here so core workflows and tests remain available before optional
    # UI dependencies are installed.
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reviewed Data"
    worksheet.append(EXPORT_COLUMNS)
    for row in rows:
        worksheet.append([row[column] for column in EXPORT_COLUMNS])

    _style_sheet(worksheet, (34, 48, 48, 14, 48))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def create_failure_export(
    failures: Sequence[Mapping[str, str]], unchecked: Sequence[Mapping[str, str]]
) -> bytes:
    """Workbook of automatic failures, plus the records no rule could judge.

    The two live on separate sheets deliberately: only the failures sheet is
    ever applied to the database, so a record that merely could not be parsed
    can never end up carrying a review it did not earn.
    """
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = FAILURE_SHEET
    worksheet.append(FAILURE_COLUMNS)
    for row in failures:
        worksheet.append([row[column] for column in FAILURE_COLUMNS])
    _style_sheet(worksheet, (12, 44, 44, 44, 34, 18, 20))

    second = workbook.create_sheet(UNCHECKED_SHEET)
    second.append(UNCHECKED_COLUMNS)
    for row in unchecked:
        second.append([row[column] for column in UNCHECKED_COLUMNS])
    _style_sheet(second, (12, 56))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
