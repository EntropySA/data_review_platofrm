#!/usr/bin/env python3
"""Record the failures found by auto_review.py in the review platform.

    python apply_review.py failures.xlsx --dry-run
    python apply_review.py failures.xlsx

Reads only the الأخطاء sheet: a record that merely could not be checked lives on
the other sheet and is never sent. Each row is matched in the database by its
source id *and* its content hash, so a row whose content has changed since the
workbook was written is reported rather than applied to something else.

Questions already failed are left alone. A question a person passed is failed,
with their review preserved in the audit trail.
"""

from __future__ import annotations

import argparse
import getpass
import json
import os
import sys
import urllib.error
import urllib.request
from pathlib import Path
from typing import Callable, Iterable, Sequence

from reporting import FAILURE_SHEET

DEFAULT_URL = "https://review-desk-api.entropy-data-review.workers.dev"
# Cloudflare's browser integrity check rejects urllib's default agent at the
# edge with a 1010, before the request ever reaches the Worker. Naming the
# tool is both honest and enough to get through.
USER_AGENT = "review-desk-apply/1.0"
CHUNK = 100
REQUIRED_COLUMNS = ("source_id", "notes", "row_hash")


class ApplyError(RuntimeError):
    pass


def read_failures(path: Path) -> list[dict]:
    """The rows to apply, carrying only the three fields the API needs.

    The content columns stay on this machine; there is no reason to send an
    instruction back to the server that already stores it.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    if FAILURE_SHEET not in workbook.sheetnames:
        raise ApplyError(f"{path} has no {FAILURE_SHEET} sheet. Was it written by auto_review.py?")
    sheet = workbook[FAILURE_SHEET]
    rows = sheet.iter_rows(values_only=True)
    header = [str(cell) if cell is not None else "" for cell in next(rows, ())]
    missing = [column for column in REQUIRED_COLUMNS if column not in header]
    if missing:
        raise ApplyError(f"{path} is missing the column(s): {', '.join(missing)}.")
    index = {column: header.index(column) for column in REQUIRED_COLUMNS}

    items = []
    for number, row in enumerate(rows, start=2):
        item = {column: str(row[position]).strip() if row[position] is not None else ""
                for column, position in index.items()}
        if not any(item.values()):
            continue
        if len(item["row_hash"]) != 64:
            raise ApplyError(f"Row {number} has no usable row_hash. Re-run auto_review.py "
                             "rather than editing the workbook by hand.")
        if not item["source_id"] or not item["notes"]:
            raise ApplyError(f"Row {number} is missing a source_id or notes.")
        items.append(item)
    workbook.close()
    return items


def chunked(items: Sequence[dict], size: int) -> Iterable[Sequence[dict]]:
    for start in range(0, len(items), size):
        yield items[start:start + size]


class Client:
    """The platform's admin API over HTTPS, using only the standard library."""

    def __init__(self, url: str, opener: Callable = urllib.request.urlopen):
        self.url = url.rstrip("/")
        self.opener = opener
        self.token = ""

    def _post(self, path: str, payload: dict) -> dict:
        request = urllib.request.Request(
            f"{self.url}{path}", method="POST",
            data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            headers={"Content-Type": "application/json", "User-Agent": USER_AGENT,
                     **({"Authorization": f"Bearer {self.token}"} if self.token else {})})
        try:
            with self.opener(request, timeout=60) as response:
                return json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as error:
            detail = error.read().decode("utf-8", "replace")
            try:
                detail = json.loads(detail).get("detail", detail)
            except ValueError:
                pass
            if "error code:" in detail:
                # Cloudflare's edge answered, not the platform, so the detail is
                # a bare code that says nothing about the request itself.
                detail += (" — this came from Cloudflare's edge, not the review platform. "
                           "The request was blocked before reaching it.")
            raise ApplyError(f"{path} failed ({error.code}): {detail}") from error
        except urllib.error.URLError as error:
            raise ApplyError(f"Could not reach {self.url}: {error.reason}") from error

    def login(self, password: str) -> None:
        response = self._post("/api/auth/login", {"password": password, "reviewer_name": ""})
        if response.get("role") != "admin":
            raise ApplyError("That password signed in as a reviewer. An admin password is needed.")
        self.token = response["token"]

    def bulk_fail(self, items: Sequence[dict]) -> dict:
        return self._post("/api/admin/reviews/bulk-fail", {"items": list(items)})


def apply_failures(client: Client, items: Sequence[dict], size: int = CHUNK,
                   progress: Callable[[int, int], None] | None = None) -> dict:
    """Send every row in batches, combining what each batch reports back."""
    totals = {"failed": 0, "overwritten": 0, "already_failed": 0, "unmatched": []}
    sent = 0
    for chunk in chunked(items, size):
        summary = client.bulk_fail(chunk)
        for key in ("failed", "overwritten", "already_failed"):
            totals[key] += int(summary.get(key, 0))
        totals["unmatched"].extend(summary.get("unmatched", []))
        sent += len(chunk)
        if progress:
            progress(sent, len(items))
    return totals


def report(totals: dict, out) -> None:
    print(f"  ✓ failed         {totals['failed']:>5}", file=out)
    print(f"  ↻ overwrote pass {totals['overwritten']:>5}", file=out)
    print(f"  – already failed {totals['already_failed']:>5}", file=out)
    unmatched = totals["unmatched"]
    print(f"  ✗ not matched    {len(unmatched):>5}", file=out)
    if unmatched:
        print(f"      source ids: {', '.join(unmatched[:40])}"
              f"{' …' if len(unmatched) > 40 else ''}", file=out)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("workbook", type=Path, help="the workbook auto_review.py wrote")
    parser.add_argument("--url", default=os.environ.get("REVIEW_URL", DEFAULT_URL),
                        help="platform base URL")
    parser.add_argument("--dry-run", action="store_true",
                        help="report what would be sent and stop before sending it")
    parser.add_argument("--yes", action="store_true", help="skip the confirmation prompt")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        items = read_failures(args.workbook)
    except (OSError, ApplyError) as exc:
        print(exc, file=sys.stderr)
        return 2

    print(f"{len(items)} failure(s) in {args.workbook}")
    print(f"target  {args.url}")
    if args.dry_run:
        print("\ndry run: nothing was sent.")
        return 0
    if not items:
        return 0
    if not args.yes:
        print("\nThis writes failures into the live database. Passed questions will be "
              "overwritten (their reviews are kept in the audit trail).")
        if input("Type yes to continue: ").strip().lower() != "yes":
            print("Nothing was sent.")
            return 1

    password = os.environ.get("REVIEW_ADMIN_PASSWORD") or getpass.getpass("Admin password: ")
    client = Client(args.url)
    try:
        client.login(password)
        totals = apply_failures(client, items,
                                progress=lambda sent, total: print(f"  sent {sent}/{total}"))
    except ApplyError as exc:
        print(exc, file=sys.stderr)
        return 1

    print()
    report(totals, sys.stdout)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
